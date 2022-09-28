import openpyxl

# # create a new blank workbook
# my_workbook = openpyxl.Workbook()
#
# # set the active worksheet
# sheet = my_workbook.active
#
# # set a title for the active worksheet
# sheet.title = "People-Data"
#
# # setting some column names
# sheet["A1"] = "NAME"
# sheet["B1"] = "TEL"
#
# # save the new workbook
# my_workbook.save(filename="s-data.xlsx")

# to read an Excel spreadsheet, first open the workbook
mgc_wb = openpyxl.load_workbook(r"..\..\..\Downloads\MGC_DATA.xlsx")

# select the default sheet from a spreadsheet.
# w_sheet = mgc_wb.active

m_sheet = mgc_wb["MGC_2"]

sm_data = {}

for row in m_sheet.iter_rows(2, values_only=True):
    subj_matter = row[0]
    sm_info = {
        "status": row[1],
        "sess_id": row[2],
        "last_date": row[3],
        "code": row[4]
    }
    sm_data[subj_matter] = sm_info

print(sm_data)
# select a specific sheet
# w_sheet2 = mgc_wb["Operations"]

# change a sheet title
# w_sheet2.title = "MGC_2"

# to create or delete sheets,use .create_sheet() and .remove(worksheet) or del wb[sheetname]:
# mgc_wb.create_sheet("Operations")

# You can also define the position to create the sheet at
# mgc_wb.create_sheet("Fury", 1)

# mgc_wb.remove(mgc_wb.get_sheet_by_name('Sheet1'))
# del mgc_wb["Operations1"]

# you can make duplicates of a sheet using copy_worksheet()
# mgc_wb.copy_worksheet(w_sheet2)

# Check the used spreadsheet space using the attribute "dimensions"
# print(w_sheet.dimensions)

# save any changes
# mgc_wb.save(r"..\..\..\Downloads\MGC_DATA.xlsx")

# get the name of the worksheets in that workbook
# print(f"WorkSheets present: {mgc_wb.sheetnames}\n")

# # get the title of the active worksheet
# print(f"Active WorkSheet Title: {w_sheet.title}\n\n")
#
# # retrive some data from the worksheet
# print(f"Value of cell A12: {w_sheet['A12'].value}\n")
#
# # alternate method to retrieve data
# print(f"Value of cell B13: {w_sheet.cell(row=13, column=2).value}\n")

# slicing the data with a combination of columns and rows to get a range of rows & cols
# some_data = w_sheet["A1:E5"]

# print(some_data)

# get all cells from a column; returns tuple
# print(f"Cells in column B: {w_sheet['B']}")
#
# # get all cells from a row; returns tuple
# print(f"Cells in row 5: {w_sheet[5]}")

"""
There are also multiple ways of using normal Python generators to go through the data. 
The main methods you can use to achieve this are:
-> .iter_rows()
-> .iter_cols()

Both methods can receive the following arguments:
min_row
max_row
min_col
max_col
These arguments are used to set boundaries for the iteration, example below:
"""
# for row_data in w_sheet.iter_rows(min_row=1, max_row=10):
#     print(row_data)

# for col_data in w_sheet.iter_cols(min_col=1, max_col=3):
#     print(col_data)

# to iterate through some data and display values
# for s_data in w_sheet.iter_rows(max_row=3, values_only=True):
#     # print(s_data)
#     for d in s_data:
#         print(f"{d}")
#     print(f"---------- End of Row -----------")

# iterate through the whole worksheet, displaying the cell name(coordinate) & cell data(value)
# for d_data in w_sheet:
#     for d in d_data:
#         print(d.coordinate, d.value )
#     print(f"---------- End of Row -----------")

# to append data to an existing:
# -> open the workbook first
s_wb = openpyxl.load_workbook("s-data.xlsx")

# # -> select the main worksheet/ select the default sheet from a spreadsheet.
s_sheet = s_wb.active

# # -> add a value to a spreadsheet
# # s_sheet["e5"] = 34_000

# add_values = (23, 122, 335, 100)

# -> edit value in a spreadsheet
# s_sheet["e5"] = 230

"""
One of the most common things you have to do when manipulating spreadsheets is adding
or removing rows and columns. The openpyxl package allows you to do that in a very
straightforward way by using the methods:
.insert_rows()
.delete_rows()
.insert_cols()
.delete_cols()
Every single one of those methods can receive two arguments:
1. idx
2. amount

NOTE:
-> When inserting new data (rows or columns), the insertion happens
before the idx parameter.
-> However, when deleting rows or columns, .delete_... deletes data starting
from the index passed as an argument.

"""
# Insert a column before the existing column 1 ("A")
# s_sheet.insert_cols(idx=1)

# Insert 2 columns before the existing column 2 ("B")
# s_sheet.insert_cols(idx=2, amount=2)

# Delete 1 column before the existing column 2 ("B")
# s_sheet.delete_cols(idx=2, amount=2)

# Insert 3 new rows before row 4
# s_sheet.insert_rows(idx=4, amount=3)

# Delete the first 6 rows after the first row (1)
# s_sheet.delete_rows(idx=2, amount=6)

# Adding a formula
# s_sheet["H4"] = "=SUM(C4:F4)"

# adding some styling
# s_sheet["B1"].font = openpyxl.styles.Font(color="00FF0000", size=21)

# to apply styling to a number of cells
# text_row = s_sheet[4]

# txt_font = openpyxl.styles.Font(color="01AA1122", size=14, bold=True)

# for cell in text_row:
#     cell.font = txt_font

# -> save the spreadsheet
s_wb.save(filename="s-data.xlsx")
